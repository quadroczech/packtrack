"""
EKO-KOM Excel report generator for PackTrack.

Loads the official EKO-KOM template (ekokom_template.xlsx),
clears data cells, fills in aggregated quarterly consumption
and returns the file as a BytesIO stream.
"""
import os
from io import BytesIO
import openpyxl

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "ekokom_template.xlsx")

# Data sheets present in the template that need to be cleared
# before filling in new values.
DATA_SHEETS = [
    "J1-1A", "J1-1B", "J1-1K",
    "J1-2A", "J1-2B", "J1-2K",
    "J1-3",  "J1-3K",
    "J2",    "J3",    "J4",
    "L",
    "O1-1A", "O1-1B", "O1-1K",
    "O1-2A", "O1-2B", "O1-2K",
    "O1-3",  "O1-3K",
    "O2",    "O3",    "O4",
    "S1",    "S2",
]

# Columns that hold user data in J1-1A / J1-1B (by letter)
# J1-1A: E-P (plastic/metal), E,G,I,K (glass/paper/wood)
# J1-1B: E-T (plastic/metal), E,G,I,K (paper/wood)
J11A_DATA_COLS = list(range(5, 17))   # E=5 … P=16
J11B_DATA_COLS = list(range(5, 21))   # E=5 … T=20

# Rows to clear in J1-1A (plastic/metal) and (glass/paper/wood)
J11A_PLASTIC_ROWS = list(range(10, 37))   # 10-36
J11A_OTHER_ROWS   = list(range(41, 53))   # 41-52

# Same for J1-1B
J11B_PLASTIC_ROWS = list(range(11, 36))   # 11-35
J11B_OTHER_ROWS   = list(range(42, 51))   # 42-50

# Simpler sheets (J2, J3, J4, O2, O3, O4) use rows 9-24 cols E-J
SIMPLE_DATA_COLS = list(range(5, 11))
SIMPLE_ROWS      = list(range(9, 25))


def _clear_sheet(ws, rows, cols):
    """Set specified cells to None, leaving formula cells untouched."""
    for r in rows:
        for c in cols:
            cell = ws.cell(row=r, column=c)
            # Skip formula cells – value starts with '='
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None


def generate_ekokom(year: int, quarter: int, cell_data: dict,
                    settings: dict) -> BytesIO:
    """
    Parameters
    ----------
    year, quarter : int
    cell_data : dict  { (sheet_name, row, col): weight_t }
    settings : dict   company settings from db.get_settings()

    Returns
    -------
    BytesIO  ready to send as file download
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # ── 1. Update Úvod (cover sheet) ────────────────────────────────────────
    # Find the intro sheet by name (handles different Unicode normalizations)
    uvod_name = next((s for s in wb.sheetnames if "vod" in s.lower()), None)
    if uvod_name is None:
        uvod_name = wb.sheetnames[0]
    uvod = wb[uvod_name]
    uvod["D4"] = quarter
    uvod["D6"] = settings.get("company_name", "")
    uvod["D8"] = settings.get("company_address", "")
    uvod["E10"] = settings.get("invoice_email", "")
    uvod["C12"] = settings.get("company_ico", "")
    uvod["F12"] = settings.get("company_dic", "")
    uvod["F14"] = settings.get("ekokom_id", "")
    uvod["D18"] = settings.get("contact_person", "")
    uvod["C19"] = settings.get("contact_phone", "")
    uvod["C21"] = settings.get("contact_email", "")

    # ── 2. Clear all data cells in every data sheet ──────────────────────────
    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if sheet_name == "J1-1A":
            _clear_sheet(ws, J11A_PLASTIC_ROWS, J11A_DATA_COLS)
            _clear_sheet(ws, J11A_OTHER_ROWS, [5, 7, 9, 11])
        elif sheet_name == "J1-1B":
            _clear_sheet(ws, J11B_PLASTIC_ROWS, J11B_DATA_COLS)
            _clear_sheet(ws, J11B_OTHER_ROWS, [5, 7, 9, 11])
        else:
            _clear_sheet(ws, SIMPLE_ROWS, SIMPLE_DATA_COLS)

    # ── 3. Write new values ──────────────────────────────────────────────────
    for (sheet_name, row, col), weight_t in cell_data.items():
        if weight_t == 0:
            continue
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cell = ws.cell(row=row, column=col)
        # Only write if it's not a formula cell
        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
            cell.value = round(weight_t, 6)

    # ── 4. Save to BytesIO ───────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_naturpack_summary(year: int, quarter: int,
                                np_data: dict, settings: dict) -> BytesIO:
    """
    Generate a simple Excel summary for NATUR-PACK data entry.
    np_data: { (appendix, material_code): weight_t }
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from reports import NATURPACK_MATERIALS, NATURPACK_GROUPS

    mat_labels = {code: label for code, label in NATURPACK_MATERIALS}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    GROUP_FILL  = PatternFill("solid", fgColor="2E74B5")
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
    THIN = Side(border_style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    company = settings.get("company_name", "")
    ico = settings.get("company_ico", "")

    for appendix, app_label in [("consumer", "Spotrebiteľské obaly (Príloha č. 2)"),
                                 ("group",    "Skupinové a prepravné obaly (Príloha č. 3)")]:
        ws = wb.create_sheet(title=app_label[:31])

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"NATUR-PACK – {app_label}"
        ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = HEADER_FILL
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:D2")
        ws["A2"] = f"{company}  |  IČO: {ico}  |  {quarter}. štvrťrok {year}"
        ws["A2"].font = Font(name="Arial", size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        # Column headers
        headers = ["Komodita", "Obalový materiál", "Hmotnosť [t]", "Poznámka"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[3].height = 22

        # Collect data grouped by commodity
        from collections import defaultdict
        groups = defaultdict(lambda: defaultdict(float))
        for (app, mat_code), weight_t in np_data.items():
            if app != appendix:
                continue
            group = NATURPACK_GROUPS.get(mat_code, "OSTATNÉ")
            groups[group][mat_code] += weight_t

        GROUP_ORDER = [
            "SKLO", "PLASTY bez PET", "PET",
            "PAPIER A LEPENKA", "NÁPOJOVÝ KARTON",
            "ŽELEZNÉ KOVY", "HLINÍK", "DREVO", "OSTATNÉ"
        ]

        row_idx = 4
        alt = False
        for group_name in GROUP_ORDER:
            if group_name not in groups:
                continue
            # Group header row
            ws.merge_cells(f"A{row_idx}:D{row_idx}")
            c = ws.cell(row=row_idx, column=1, value=group_name)
            c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            c.fill = GROUP_FILL
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = BORDER
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

            for mat_code, weight_t in groups[group_name].items():
                fill = ALT_FILL if alt else WHITE_FILL
                ws.cell(row=row_idx, column=1, value="").fill = fill
                ws.cell(row=row_idx, column=1).border = BORDER
                label_cell = ws.cell(row=row_idx, column=2,
                                     value=mat_labels.get(mat_code, mat_code))
                label_cell.font = Font(name="Arial", size=10)
                label_cell.fill = fill
                label_cell.border = BORDER
                weight_cell = ws.cell(row=row_idx, column=3, value=round(weight_t, 3))
                weight_cell.font = Font(name="Arial", size=10, bold=True)
                weight_cell.fill = fill
                weight_cell.border = BORDER
                weight_cell.number_format = "0.000"
                ws.cell(row=row_idx, column=4, value="Dovoz (Cp)").fill = fill
                ws.cell(row=row_idx, column=4).font = Font(name="Arial", size=9,
                                                            color="888888", italic=True)
                ws.cell(row=row_idx, column=4).border = BORDER
                ws.row_dimensions[row_idx].height = 16
                alt = not alt
                row_idx += 1

        # Total row
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        tot_cell = ws.cell(row=row_idx, column=1, value="CELKOM")
        tot_cell.font = Font(name="Arial", bold=True, size=11)
        tot_cell.fill = HEADER_FILL
        tot_cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        tot_cell.border = BORDER

        total = sum(v for (app, _), v in np_data.items() if app == appendix)
        tc = ws.cell(row=row_idx, column=3, value=round(total, 3))
        tc.font = Font(name="Arial", bold=True, size=11)
        tc.fill = HEADER_FILL
        tc.font = Font(name="Arial", bold=True, color="FFFFFF")
        tc.number_format = "0.000"
        tc.border = BORDER
        ws.cell(row=row_idx, column=4).fill = HEADER_FILL
        ws.cell(row=row_idx, column=4).border = BORDER
        ws.row_dimensions[row_idx].height = 20

        # Column widths
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
